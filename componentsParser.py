import csv
import os
from openpyxl import load_workbook, Workbook
import utilityFiles as Ufiles
import cParser
import cUsage


def update_table(xlsx_file, component, comp_functions, total_functions, num_vars, comp_vars, total_vars, num_headers):
    try:
        wb = load_workbook(xlsx_file)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["CompName", "NrUpdates", "PolyCyclomatic", "NrLines", "GlobalVars",
                   "NrHeaders", "ConnectedComp", "nr FuncCalls", "nr VarRead", "nr VarWrite",
                   "ArchComp", "Complexity"])

    row = ws.max_row + 1 if ws.max_row > 1 or ws.cell(1, 1).value else 2

    ws.cell(row=row, column=1, value=component)
    ws.cell(row=row, column=5, value=num_vars)
    ws.cell(row=row, column=6, value=num_headers)
    ws.cell(row=row, column=7, value="Total")
    ws.cell(row=row, column=8, value=total_functions)
    ws.cell(row=row, column=9, value=total_vars)

    remaining_comp_vars = dict(comp_vars)

    for i, (comp_name, num_functions) in enumerate(comp_functions.items()):
        ws.cell(row=row + i + 1, column=7, value=comp_name)
        ws.cell(row=row + i + 1, column=8, value=num_functions)

        if comp_name in remaining_comp_vars:
            ws.cell(row=row + i + 1, column=9, value=remaining_comp_vars[comp_name])
            del remaining_comp_vars[comp_name]

    new_row = ws.max_row + 1
    for i, (comp_name, num_vars) in enumerate(remaining_comp_vars.items()):
        ws.cell(row=new_row + i, column=7, value=comp_name)
        ws.cell(row=new_row + i, column=9, value=num_vars)

    wb.save(xlsx_file)

def csv_parser(csv_file, xlsx_file):
    all_functions_files = {}
    all_variables_files = {}
    all_headers_components = {}
    all_nr_GlobalVariables = {}


    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:

            path = row[1]
            if os.path.exists(path):
                component = row[0]
                #parse comp folder and get paths to the valid files (exclude tempalte/generated/..)
                c_files = Ufiles.check_valid_c(path)
                h_files = Ufiles.check_valid_h(path)

                compLines = 0
                nr_GlobalVariables = 0

                headers_components = Ufiles.get_comp_headers(path, component)
                all_headers_components.update(headers_components)
                #get global variable definitions from component
                for h_file in h_files:
                    #extract defines
                    #defines_filesH = cParser.extract_c_defines(h_file, component)
                    # copo deactivated for now, defines should not be counted as global variables
                    # all_variables_files.update(defines_filesH)
                    #get number of lines excluding coments, empty lines
                    compLines += len(Ufiles.JustValidLines(h_file))
                #get global variables and functions defined in component
                for c_file in c_files:
                    # get all defines from the c file
                    #defines_filesC = cParser.extract_c_defines(h_file, component)
                    #copo deactivated for now, defines should not be counted as global variables
                    # all_variables_files.update(defines_filesC)

                    #get all variables from the c file
                    variables_filesC = (cParser.extract_c_variables(c_file, component))
                    nr_GlobalVariables += len(variables_filesC)
                    all_variables_files.update(variables_filesC)

                    # get all functions defined in component
                    functions_files = cParser.extract_c_functions(c_file, component)
                    all_functions_files.update(functions_files)

                    #get number of lines excluding coments, empty lines
                    compLines += len(Ufiles.JustValidLines(c_file))

                all_nr_GlobalVariables[component] = nr_GlobalVariables
                print(f'---------------------\n{component}\n')
                # print(path)
                # print("c_files:" + str(c_files))
                # print("h_files:" + str(h_files))
                # print("functions_in_component:" + str(functions_files))
                # print("headers_in_component:" + str(headers_components))
                # print("variables_filesC:" + str(variables_filesC))
                print("nr_GlobalVariables:" + str(nr_GlobalVariables))
                print("Valid_Line_numbers:" + str(compLines))

            else:
                print("component not found at path: " + path)
    print(all_nr_GlobalVariables)
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)  # Skip header
        for row in reader:
            path = row[1]
            component = row[0]
            # get the number of global variables defined inside the component
            num_variables = all_nr_GlobalVariables.get(component, 0)
            try:
                #get all the fucntion calls to other modules
                components_functions, total_functions = cUsage.count_external_function_usage(path, all_functions_files, component)
            except:
                print("problem for comp1:" + component)
            try:
                components_variables, total_variables = cUsage.count_external_variable_usage(path, all_variables_files, component)
            except:
                print("problem for comp2:" + component)
            try:
                #get number of headers used from other components
                num_headers = cUsage.nr_external_headers(path, all_headers_components, component)
            except:
                print("problem for comp3:" + component)
            try:
                update_table(xlsx_file, component, components_functions, total_functions, num_variables, components_variables, total_variables, num_headers)
            except:
                print("excel update crush")
csv_parser("updated_locations.csv", "Complexity.xlsx")