import os
import re
import openpyxl
from openpyxl import Workbook

def JustValidLines(file_path):
    """
    Reads a C file and removes all comments, unnecessary new lines, and preserves valid code formatting.

    Args:
        file_path (str): Path to the C file.

    Returns:
        list: A list of valid code lines with comments removed.
    """
    valid_lines = []
    inside_block_comment = False
    previous_line_ends_with_comma = False

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            stripped_line = line.strip()

            # Check for block comments /* ... */
            if "/*" in stripped_line:
                inside_block_comment = True
                if "*/" in stripped_line:  # Handle single-line block comments
                    inside_block_comment = False
                continue

            if inside_block_comment:
                if "*/" in stripped_line:
                    inside_block_comment = False
                continue

            # Remove inline comments (// ...)
            line = re.sub(r'//.*', '', line).strip()

            # Skip empty lines unless needed
            if not line:
                continue

            # If previous line ended with ',', append this line to the previous line
            if previous_line_ends_with_comma:
                valid_lines[-1] += " " + line  # Merge with previous line
            else:
                valid_lines.append(line)

            # Update the flag for next iteration
            previous_line_ends_with_comma = line.endswith(",")

    return valid_lines



def check_valid_c(path):
    valid_paths = []
    for root, dirs, files in os.walk(path):
        for file in files:
            file_path = os.path.join(root, file)
            if(file_path.__contains__("autosar_rtw" or "template")):
                file_ignored = file_path
            else:
                if(file_path.endswith(".c")):
                   valid_paths.append(file_path)
    return valid_paths

def check_valid_h(path):
    valid_paths = []
    for root, dirs, files in os.walk(path):
        for file in files:
            file_path = os.path.join(root, file)
            if(file_path.__contains__("autosar_rtw" or "template")):
                file_ignored = file_path
            else:
                if(file_path.endswith(".h")):
                   valid_paths.append(file_path)
    return valid_paths

def find_headers(path, component):
    headers_components = {}
    for root, dirs, files in os.walk(path):
        for name in files:
            if name.endswith(".h"):
                headers_components[name] = component
    return headers_components


def extract_function_name(line):
    match = re.search(r"\b\w+\s+\**(\w+)\s*\(", line)
    return match.group(1) if match else None


def extract_c_functions(file_path, comp):
    """
    Extracts all function definitions from a C file, including macro-based functions.

    Args:
        file_path (str): Path to the C file.
        comp (str): Component name.

    Returns:
        dict: A dictionary mapping function names to their component.
    """
    # Improved regex to match FUNC(...) macros and standard function definitions
    function_pattern = re.compile(
        r'^\s*(?:FUNC\s*\(\s*([\w\s,\*]+)\s*,\s*[\w_]+\s*\)|([\w\s\*]+))\s+([a-zA-Z_]\w*)\s*\(([^)]*)\)\s*'
    )

    function_names = {}
    functions = []
    in_function_declaration = False
    function_header = ""

    clean_lines = JustValidLines(file_path)  # Preprocess file

    for line in clean_lines:
        line = line.strip()
        if comp == "CDSM":
            print(line)
        # Skip preprocessor macros and variable definitions
        if line.startswith("#") or line.startswith("VAR("):
            continue

        # Handle multi-line function definitions
        if in_function_declaration:
            function_header += " " + line
            if "{" in line:  # End of function declaration
                function_header = function_header.replace("{", "").strip()
                in_function_declaration = False
                match = function_pattern.match(function_header)
                if match:
                    return_type = match.group(1) or match.group(2)
                    function_name = match.group(3)
                    function_names[function_name] = comp
                    functions.append((return_type.strip(), function_name.strip()))
            continue

        # Match function definitions (without '{' yet)
        match = function_pattern.match(line)
        if match:
            in_function_declaration = True
            function_header = line
            continue

        # Check if this is the opening '{' of a function
        if in_function_declaration and "{" in line:
            in_function_declaration = False
            function_header = function_header.replace("{", "").strip()
            match = function_pattern.match(function_header)
            if match:
                return_type = match.group(1) or match.group(2)
                function_name = match.group(3)
                function_names[function_name] = comp
                functions.append((return_type.strip(), function_name.strip()))

    print("Extracted functions:", function_names)
    return function_names


def extrage_nume_variabila(linie):
    # Căutăm numele variabilei după tipul de date, urmat fie de '=' fie de ';'
    match = re.search(r'\b(?:\w+\s+)+(\w+)\s*(?:=|;)', linie)
    if match:
        return match.group(1)  # Extragem doar numele variabilei
    return None


def extract_define_name(line):
    match = re.search(r"#define\s+(\w+)", line)
    if match:
        return match.group(1)
    return None

def find_variables(input_file, comp):
    with open(input_file, 'r') as file:
        lines = file.readlines()
        inside_function = 0
        variables = []
        variable_directories = {}

        for line in lines:
            if '{' in line:
                inside_function += 1
            if '}' in line:
                inside_function -= 1
            # Verificăm dacă linia conține un tip de date și o variabilă
            if not inside_function and re.search(r'\b(?:\w+\s+)+\w+\s*(?:=|;)', line) and not '(' in line:
                variable = extrage_nume_variabila(line.strip())
                if variable:
                    variables.append(variable)
                    variable_directories[variable] = comp
            if line.startswith("#"):
                define = extract_define_name(line)
                if define:
                    variables.append(define)
                    variable_directories[define] = comp

    return variables, variable_directories

def find_defines(input_file):
    with open(input_file, 'r') as file:
        lines = file.readlines()
        inside_function = 0
        defines_directories = {}

        for line in lines:
            if '{' in line:
                inside_function += 1
            if '}' in line:
                inside_function -= 1
            if line.startswith("#"):
                define = extract_define_name(line)
                if define:
                    defines_directories[define] = input_file

    return defines_directories

def find_calls(input_file, function):
    with (open(input_file, 'r') as file):
        lines = file.readlines()
        inside_function = 0
        called_functions = {}
        called_functions[function] = []
        found = False
        for i in range(len(lines) - 1):
            line = lines[i]
            if function in line and '{' in lines[i+1]:
                found = True

            if found:
                if '{' in line:
                    inside_function += 1

                if '}' in line:
                    inside_function -= 1

                    if inside_function == 0:
                        break
            if inside_function > 0:
                if '(' in line and ');' in line:
                    equal_idx = line.find('=')
                    paren_idx = line.find('(')

                    if equal_idx != -1:
                        if line[equal_idx+1:paren_idx].strip():
                            func_call = line[equal_idx+1:].strip()
                        else:
                            continue
                    else:
                        func_call = lines[i].strip()



                    #print(f"{input_file}: {function}: {func_call}")
                    called_functions[function].append(func_call)

    return called_functions

# def plot_graph(G):
#     pos = nx.spring_layout(G)  # Layout-ul graficului
#
#     # Extragem coordonatele nodurilor
#     x_nodes = [pos[node][0] for node in G.nodes()]
#     y_nodes = [pos[node][1] for node in G.nodes()]
#
#     # Extragem coordonatele muchiilor
#     edge_x = []
#     edge_y = []
#     for edge in G.edges():
#         x0, y0 = pos[edge[0]]
#         x1, y1 = pos[edge[1]]
#         edge_x.extend([x0, x1, None])
#         edge_y.extend([y0, y1, None])
#
#     # Muchiile graficului
#     edge_trace = go.Scatter(
#         x=edge_x, y=edge_y,
#         line=dict(width=1, color='black'),
#         hoverinfo='none',
#         mode='lines')
#
#     # Nodurile graficului
#     node_trace = go.Scatter(
#         x=x_nodes, y=y_nodes,
#         mode='markers+text',
#         text=list(G.nodes()),
#         textposition="bottom center",
#         hoverinfo='text',
#         marker=dict(
#             showscale=True,
#             colorscale='YlGnBu',
#             size=10,
#             colorbar=dict(
#                 thickness=15,
#                 title='Funcții',
#                 xanchor='left',
#                 titleside='right'
#             ),
#             line_width=2))
#
#     # Creăm figura și adăugăm layout-ul
#     fig = go.Figure(data=[edge_trace, node_trace],
#                     layout=go.Layout(
#                         title='Grafic de Funcții Apelate',
#                         titlefont_size=16,
#                         showlegend=False,
#                         hovermode='closest',
#                         margin=dict(b=0, l=0, r=0, t=0),
#                         xaxis=dict(showgrid=False, zeroline=False),
#                         yaxis=dict(showgrid=False, zeroline=False))
#                     )
#
#
#     fig.show()

def create_excell_output(output_file, file, function, called_function, file_of_called_function):

        try:
            workbook = openpyxl.load_workbook(output_file)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.cell(row=1, column=1, value="File")
            sheet.cell(row=1, column=2, value="Function")
            sheet.cell(row=1, column=3, value="Called Function")
            sheet.cell(row=1, column=4, value="File Of Called Function")

        row = sheet.max_row + 1 if sheet.max_row > 1 or sheet.cell(1, 1).value else 1

        # Scrie variabilele în coloane pe rândul gol găsit
        sheet.cell(row=row, column=1, value=file)
        sheet.cell(row=row, column=2, value=function)
        sheet.cell(row=row, column=3, value=called_function)
        sheet.cell(row=row, column=4, value=file_of_called_function)

        workbook.save(output_file)


def create_excell_output_variables(output_file, file, used_variable, file_of_used_variable):
    try:
        workbook = openpyxl.load_workbook(output_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value="File")
        sheet.cell(row=1, column=2, value="Used Variable")
        sheet.cell(row=1, column=3, value="File Of Used Variable")

    row = sheet.max_row + 1 if sheet.max_row > 1 or sheet.cell(1, 1).value else 1

    # Scrie variabilele în coloane pe rândul gol găsit
    sheet.cell(row=row, column=1, value=file)
    sheet.cell(row=row, column=2, value=used_variable)
    sheet.cell(row=row, column=3, value=file_of_used_variable)

    workbook.save(output_file)

def process_c_files(directory):

    all_variables = []
    all_variables_directories = {}
    all_functions = []
    all_functions_directories = {}
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.c'):
                full_path = os.path.join(root, file)
                functions, function_directories = find_functions(full_path)
                all_functions.extend(functions)
                all_functions_directories.update(function_directories)
                variables, variable_directories = find_variables(full_path)
                all_variables.extend(variables)
                all_variables_directories.update(variable_directories)

    seen_entries = set()

    # Căutăm utilizările variabilelor
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            with open(file_path, "r", encoding="utf-8") as f:
                lines = f.readlines()
                for line in lines:
                    for variable in all_variables:
                        if variable in line:
                            file_name = all_variables_directories[variable]
                            if file_path != file_name:
                                entry = (file_path, variable, file_name)
                                # Verificăm dacă intrarea a fost deja adăugată
                                if entry not in seen_entries:
                                    seen_entries.add(entry)
                                    create_excell_output_variables("Variable Uses.xlsx", file_path, variable, file_name)


    for function in all_functions:
        file_name = all_functions_directories.get(function)
        #print(file_name)
        called_functions = find_calls(file_name, function)
        for called_func in called_functions:
            file_of_called_func = all_functions_directories.get(called_func)
            create_excell_output("Function Calls.xlsx", file_name, function, called_func, file_of_called_func)



directory = r'./InputFiles'
process_c_files(directory)
